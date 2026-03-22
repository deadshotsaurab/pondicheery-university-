# fill_output_results.py
# Fills both sheets of Output_results.xlsx using GMM vocabulary + readability formulas

import os, re, math, csv
import pandas as pd
import openpyxl

# --- AUTO-DETECT FOLDERS ---
def find_folder(candidates):
    for p in candidates:
        if os.path.isdir(p):
            return p
    raise FileNotFoundError("Folder not found: " + str(candidates))

JUDG_DIR = find_folder(["data/IN-Ext_judgement", "IN-Ext_judgement"])
GT_DIR   = find_folder(["data/A1_Ground truth", "data/summary/A1_Ground truth", "A1_Ground truth"])
LK_DIR   = find_folder(["data/A2_LocknKey", "data/summary/A2_LocknKey", "A2_LocknKey"])

XLSX_PATH = next(
    (f for f in ["Output_results.xlsx", "Output results.xlsx"] if os.path.exists(f)), None
)
if XLSX_PATH is None:
    raise FileNotFoundError("Cannot find Output_results.xlsx in this folder.")

print("=" * 60)
print("FILLING Output_results.xlsx - BOTH SHEETS")
print("=" * 60)
print("  Judgements  :", JUDG_DIR)
print("  Ground Truth:", GT_DIR)
print("  LockNKey    :", LK_DIR)
print("  Excel file  :", XLSX_PATH)

# --- READ TEXT FILE ---
def read(folder, fname):
    return open(os.path.join(folder, fname), encoding="utf-8", errors="ignore").read()

# --- SYLLABLE COUNTER ---
VOWELS = set("aeiou")

def syllables(word):
    w = word.lower()
    c, prev = 0, False
    for ch in w:
        v = ch in VOWELS
        if v and not prev:
            c += 1
        prev = v
    if w.endswith("e") and c > 1:
        c -= 1
    return max(1, c)

# --- READABILITY METRICS ---
def parse(text):
    sents = [s for s in re.split(r'(?<=[.!?])\s+', text.strip()) if s.strip()]
    words = re.findall(r"[a-zA-Z]+", text)
    return sents, words

def flesch(text):
    sents, words = parse(text)
    ns = max(1, len(sents))
    nw = max(1, len(words))
    nsyl = sum(syllables(w) for w in words)
    return round(206.835 - 1.015 * (nw / ns) - 84.6 * (nsyl / nw), 4)

def fog(text):
    sents, words = parse(text)
    ns = max(1, len(sents))
    nw = max(1, len(words))
    nc = sum(1 for w in words if syllables(w) >= 3)
    return round(0.4 * ((nw / ns) + 100 * (nc / nw)), 4)

def smog(text):
    sents, words = parse(text)
    ns = max(1, len(sents))
    nc = sum(1 for w in words if syllables(w) >= 3)
    return round(3 + math.sqrt(nc * 30 / ns), 4)

# --- LOAD GMM VOCABULARY ---
print("\nLoading GMM vocabulary from output/ folder...")

layman_v  = set(pd.read_csv(os.path.join("output", "layman_vocabulary.csv"))["word"].str.lower())
student_v = set(pd.read_csv(os.path.join("output", "student_vocabulary.csv"))["word"].str.lower())
prof_v    = set(pd.read_csv(os.path.join("output", "professional_vocabulary.csv"))["word"].str.lower())

print("  LAYMAN       :", len(layman_v), "words")
print("  STUDENT      :", len(student_v), "words")
print("  PROFESSIONAL :", len(prof_v), "words")

# --- GMM SCORING ---
# Each score = count of matching words / TOTAL document tokens * 100
# So the 3 scores are independent (do NOT sum to 100)
# Proposed Avg = (Layman + Student + Professional) / 3  --> meaningful unique value per doc
def gmm_scores(text):
    tokens = re.findall(r"[a-z]{3,}", text.lower())
    total = max(1, len(tokens))

    lc = sum(1 for w in tokens if w in layman_v)
    sc = sum(1 for w in tokens if w in student_v)
    pc = sum(1 for w in tokens if w in prof_v)

    lay = round(100 * lc / total, 4)
    stu = round(100 * sc / total, 4)
    pro = round(100 * pc / total, 4)

    # Proposed Metric Avg = (Layman + BuddingLawyer + Advocate) / 3
    proposed = round((lay + stu + pro) / 3, 4)

    return lay, stu, pro, proposed

# --- COMPUTE ALL SCORES ---
case_files = sorted(os.listdir(JUDG_DIR))
print("\nComputing scores for", len(case_files), "cases...\n")

all_scores = {}

for i, fname in enumerate(case_files):
    n = i + 1
    all_scores[n] = {}

    for label, folder in [("OJ", JUDG_DIR), ("GT", GT_DIR), ("LK", LK_DIR)]:
        text = read(folder, fname)
        lay, stu, pro, avg = gmm_scores(text)
        all_scores[n][label] = {
            "flesch":   flesch(text),
            "fog":      fog(text),
            "smog":     smog(text),
            "proposed": avg,
            "layman":   lay,
            "student":  stu,
            "advocate": pro,
        }

    s = all_scores[n]
    print("  Case", n, "(" + fname.replace(".txt", "") + ")")
    print("    Flesch  OJ=", s["OJ"]["flesch"], " GT=", s["GT"]["flesch"], " LK=", s["LK"]["flesch"])
    print("    Fog     OJ=", s["OJ"]["fog"],    " GT=", s["GT"]["fog"],    " LK=", s["LK"]["fog"])
    print("    SMOG    OJ=", s["OJ"]["smog"],   " GT=", s["GT"]["smog"],   " LK=", s["LK"]["smog"])
    print("    Prop    OJ=", s["OJ"]["proposed"]," GT=", s["GT"]["proposed"]," LK=", s["LK"]["proposed"])

# --- FILL SHEET 1: Overall Comparion ---
# Case N starts at row = 3 + (N-1)*4
# row+0=Flesch, row+1=Fog, row+2=SMOG, row+3=Proposed
# Col 3=OJ, Col 6=GT, Col 9=LK

print("\nFilling Sheet 1: Overall Comparion...")
wb  = openpyxl.load_workbook(XLSX_PATH)
ws1 = wb["Overall Comparion"]

METRICS = ["flesch", "fog", "smog", "proposed"]

for n, scores in all_scores.items():
    base = 3 + (n - 1) * 4
    for offset, metric in enumerate(METRICS):
        row = base + offset
        ws1.cell(row=row, column=3).value = scores["OJ"][metric]
        ws1.cell(row=row, column=6).value = scores["GT"][metric]
        ws1.cell(row=row, column=9).value = scores["LK"][metric]

print("  OK -", len(all_scores), "cases x 4 metrics x 3 doc types filled")

# --- FILL SHEET 2: Audience-aware capability ---
# Case N starts at row = 4 + (N-1)*3
# row+0=OJ, row+1=GT, row+2=LK
# Col 4=Layman, Col 5=Student, Col 6=Advocate

print("Filling Sheet 2: Audience-aware capability...")
ws2 = wb["Audience-aware capability"]

for n, scores in all_scores.items():
    base = 4 + (n - 1) * 3
    for label, offset in [("OJ", 0), ("GT", 1), ("LK", 2)]:
        row = base + offset
        ws2.cell(row=row, column=4).value = scores[label]["layman"]
        ws2.cell(row=row, column=5).value = scores[label]["student"]
        ws2.cell(row=row, column=6).value = scores[label]["advocate"]

print("  OK -", len(all_scores), "cases x 3 doc types x 3 audience scores filled")

# --- SAVE EXCEL ---
wb.save(XLSX_PATH)
print("\nDONE - Saved:", XLSX_PATH)
print("  Sheet 1 - Overall Comparion        ->", len(all_scores) * 4 * 3, "cells filled")
print("  Sheet 2 - Audience-aware capability ->", len(all_scores) * 3 * 3, "cells filled")

# --- SAVE CSV BACKUPS ---
os.makedirs("output", exist_ok=True)

rows_s1, rows_s2 = [], []
for n, scores in all_scores.items():
    for metric in METRICS:
        rows_s1.append({
            "case_num": n,
            "metric": metric,
            "original_judgement": scores["OJ"][metric],
            "ground_truth": scores["GT"][metric],
            "locknkey": scores["LK"][metric],
        })
    for label, doc_name in [("OJ", "Original Judgement"), ("GT", "Ground Truth"), ("LK", "LockNKey")]:
        rows_s2.append({
            "case_num": n,
            "doc_type": doc_name,
            "layman": scores[label]["layman"],
            "budding_lawyer": scores[label]["student"],
            "advocate": scores[label]["advocate"],
        })

with open("output/overall_comparison.csv", "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=list(rows_s1[0].keys()))
    w.writeheader()
    w.writerows(rows_s1)

with open("output/audience_capability.csv", "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=list(rows_s2[0].keys()))
    w.writeheader()
    w.writerows(rows_s2)

print("  CSV backup 1 -> output/overall_comparison.csv")
print("  CSV backup 2 -> output/audience_capability.csv")