import os
import re
import math
import pandas as pd
from openpyxl import load_workbook

# -----------------------------
# PATH CONFIGURATION
# -----------------------------

BASE_DIR = "data"

OJ_DIR = os.path.join(BASE_DIR, "IN-Ext_judgement")
GT_DIR = os.path.join(BASE_DIR, "A1_Ground truth")
LK_DIR = os.path.join(BASE_DIR, "A2_LocknKey")

EXCEL_FILE = os.path.join(BASE_DIR, "Output results.xlsx")

LAYMAN_FILE = "layman_vocabulary.csv"
STUDENT_FILE = "student_vocabulary.csv"
PROF_FILE = "professional_vocabulary.csv"


# -----------------------------
# LOAD VOCABULARY
# -----------------------------

layman_vocab = set(pd.read_csv(LAYMAN_FILE)["word"].str.lower())
student_vocab = set(pd.read_csv(STUDENT_FILE)["word"].str.lower())
prof_vocab = set(pd.read_csv(PROF_FILE)["word"].str.lower())


# -----------------------------
# TEXT UTILITIES
# -----------------------------

def tokenize(text):
    return re.findall(r"[a-zA-Z]+", text.lower())


def sentence_count(text):
    sentences = re.split(r"[.!?]+", text)
    return max(1, len([s for s in sentences if s.strip()]))


def syllable_count(word):
    word = word.lower()
    vowels = "aeiouy"
    count = 0
    prev = False

    for c in word:
        if c in vowels:
            if not prev:
                count += 1
            prev = True
        else:
            prev = False

    if word.endswith("e"):
        count = max(1, count - 1)

    return max(1, count)


# -----------------------------
# READABILITY METRICS
# -----------------------------

def flesch(text):

    words = tokenize(text)
    wc = len(words)
    sc = sentence_count(text)

    syllables = sum(syllable_count(w) for w in words)

    if wc == 0:
        return 0

    return 206.835 - 1.015*(wc/sc) - 84.6*(syllables/wc)


def gunning_fog(text):

    words = tokenize(text)
    wc = len(words)
    sc = sentence_count(text)

    complex_words = sum(1 for w in words if syllable_count(w) >= 3)

    if wc == 0:
        return 0

    return 0.4*((wc/sc) + 100*(complex_words/wc))


def smog(text):

    words = tokenize(text)
    sc = sentence_count(text)

    complex_words = sum(1 for w in words if syllable_count(w) >= 3)

    if sc == 0:
        return 0

    return 1.043 * math.sqrt(complex_words * 30 / sc) + 3.1291


# -----------------------------
# GMM VOCABULARY AUDIENCE SCORE
# -----------------------------

def audience_scores(text):

    words = tokenize(text)
    total = len(words)

    if total == 0:
        return 0,0,0,0

    lay = sum(1 for w in words if w in layman_vocab)
    stu = sum(1 for w in words if w in student_vocab)
    pro = sum(1 for w in words if w in prof_vocab)

    lay_p = (lay/total)*100
    stu_p = (stu/total)*100
    pro_p = (pro/total)*100

    avg = (lay_p + stu_p + pro_p)/3

    return lay_p, stu_p, pro_p, avg


# -----------------------------
# LOAD EXCEL
# -----------------------------

wb = load_workbook(EXCEL_FILE)

sheet1 = wb[wb.sheetnames[0]]
sheet2 = wb[wb.sheetnames[1]]


# -----------------------------
# PROCESS FILES
# -----------------------------

files = sorted(os.listdir(OJ_DIR))

row1 = 3
row2 = 3

print("Processing", len(files), "cases...\n")

for fname in files:

    oj_text = open(os.path.join(OJ_DIR, fname), encoding="utf-8", errors="ignore").read()
    gt_text = open(os.path.join(GT_DIR, fname), encoding="utf-8", errors="ignore").read()
    lk_text = open(os.path.join(LK_DIR, fname), encoding="utf-8", errors="ignore").read()

    # Readability
    f_oj = flesch(oj_text)
    f_gt = flesch(gt_text)
    f_lk = flesch(lk_text)

    fog_oj = gunning_fog(oj_text)
    fog_gt = gunning_fog(gt_text)
    fog_lk = gunning_fog(lk_text)

    smog_oj = smog(oj_text)
    smog_gt = smog(gt_text)
    smog_lk = smog(lk_text)

    # Audience
    lay_oj, stu_oj, pro_oj, avg_oj = audience_scores(oj_text)
    lay_gt, stu_gt, pro_gt, avg_gt = audience_scores(gt_text)
    lay_lk, stu_lk, pro_lk, avg_lk = audience_scores(lk_text)

    # -----------------------------
    # SHEET 1 (Overall Comparison)
    # -----------------------------

    sheet1[f"C{row1}"] = f_oj
    sheet1[f"D{row1}"] = f_gt
    sheet1[f"E{row1}"] = f_lk

    sheet1[f"C{row1+1}"] = fog_oj
    sheet1[f"D{row1+1}"] = fog_gt
    sheet1[f"E{row1+1}"] = fog_lk

    sheet1[f"C{row1+2}"] = smog_oj
    sheet1[f"D{row1+2}"] = smog_gt
    sheet1[f"E{row1+2}"] = smog_lk

    sheet1[f"C{row1+3}"] = avg_oj
    sheet1[f"D{row1+3}"] = avg_gt
    sheet1[f"E{row1+3}"] = avg_lk

    row1 += 4

    # -----------------------------
    # SHEET 2 (Audience Awareness)
    # -----------------------------

    sheet2[f"D{row2}"] = lay_oj
    sheet2[f"E{row2}"] = stu_oj
    sheet2[f"F{row2}"] = pro_oj

    sheet2[f"D{row2+1}"] = lay_gt
    sheet2[f"E{row2+1}"] = stu_gt
    sheet2[f"F{row2+1}"] = pro_gt

    sheet2[f"D{row2+2}"] = lay_lk
    sheet2[f"E{row2+2}"] = stu_lk
    sheet2[f"F{row2+2}"] = pro_lk

    row2 += 3


# -----------------------------
# SAVE FILE
# -----------------------------

OUTPUT_FILE = "Output_results_FILLED.xlsx"

wb.save(OUTPUT_FILE)

print("\nValidation completed successfully.")
print("Saved:", OUTPUT_FILE)