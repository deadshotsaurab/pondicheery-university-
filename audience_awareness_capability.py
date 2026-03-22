"""
audience_awareness_capability.py
=================================
Computes Audience-Aware Capability scores for all 50 Supreme Court cases
and writes them DIRECTLY into Sheet 2 "Audience-aware capability" of
Output results.xlsx — no manual work needed.

Run:
    python audience_awareness_capability.py

Requires:
    output/layman_vocabulary.csv
    output/student_vocabulary.csv
    output/professional_vocabulary.csv
    Output results.xlsx

Output:
    Output results.xlsx  (Sheet 2 filled automatically)
    output/audience_awareness_scores.csv
"""

import os
import re
import csv
import pandas as pd
import openpyxl

# ── PATHS (auto-detects your folder structure) ─────────────────────
def find_folder(candidates):
    for p in candidates:
        if os.path.isdir(p):
            return p
    raise FileNotFoundError(f"Could not find any of: {candidates}")

JUDG_DIR = find_folder([
    os.path.join("data", "IN-Ext_judgement"),
    "IN-Ext_judgement",
])
GT_DIR = find_folder([
    os.path.join("data", "summary", "A1_Ground truth"),
    os.path.join("data", "A1_Ground truth"),
    os.path.join("summary", "A1_Ground truth"),
    "A1_Ground truth",
])
LK_DIR = find_folder([
    os.path.join("data", "summary", "A2_LocknKey"),
    os.path.join("data", "A2_LocknKey"),
    os.path.join("summary", "A2_LocknKey"),
    "A2_LocknKey",
])
VOCAB_DIR = "output"
XLSX_PATH = "Output results.xlsx"
OUT_CSV   = os.path.join("output", "audience_awareness_scores.csv")
os.makedirs("output", exist_ok=True)

print(f"  JUDG folder : {JUDG_DIR}")
print(f"  GT folder   : {GT_DIR}")
print(f"  LK folder   : {LK_DIR}\n")

# ── SHEET 2 COLUMN MAP ─────────────────────────────────────────────
# "Audience-aware capability" sheet layout:
#   Row 4  = Case 1 → Original Judgement
#   Row 5  = Case 1 → Ground Truth Summary
#   Row 6  = Case 1 → LockNKey Summary
#   Row 7  = Case 2 → Original Judgement  ... and so on
#
#   Col D (4) = Layman Score %
#   Col E (5) = Budding Lawyer Score %
#   Col F (6) = Advocate Score %

COL_LAYMAN  = 4   # Column D
COL_STUDENT = 5   # Column E
COL_ADVOCATE= 6   # Column F

def sheet2_row(case_num, doc_offset):
    """
    case_num   : 1-50
    doc_offset : 0=Original Judgement, 1=Ground Truth, 2=LockNKey
    """
    return 4 + (case_num - 1) * 3 + doc_offset

# ── LOAD GMM VOCABULARY ────────────────────────────────────────────
print("=" * 60)
print("AUDIENCE-AWARE CAPABILITY — FILLING EXCEL SHEET 2")
print("=" * 60)
print("\nLoading GMM vocabulary...")

layman_vocab  = set(pd.read_csv(os.path.join(VOCAB_DIR, "layman_vocabulary.csv"))["word"].str.lower())
student_vocab = set(pd.read_csv(os.path.join(VOCAB_DIR, "student_vocabulary.csv"))["word"].str.lower())
prof_vocab    = set(pd.read_csv(os.path.join(VOCAB_DIR, "professional_vocabulary.csv"))["word"].str.lower())

print(f"  LAYMAN       : {len(layman_vocab)} words")
print(f"  STUDENT      : {len(student_vocab)} words  (Budding Lawyer)")
print(f"  PROFESSIONAL : {len(prof_vocab)} words  (Advocate)")

# ── SCORE FUNCTION ─────────────────────────────────────────────────
def audience_scores(text):
    tokens  = re.findall(r"[a-z]{3,}", text.lower())
    lay_c   = sum(1 for w in tokens if w in layman_vocab)
    stu_c   = sum(1 for w in tokens if w in student_vocab)
    pro_c   = sum(1 for w in tokens if w in prof_vocab)
    total   = max(1, lay_c + stu_c + pro_c)
    return (
        round(100 * lay_c / total, 4),
        round(100 * stu_c / total, 4),
        round(100 * pro_c / total, 4),
    )

# ── COMPUTE ALL SCORES ─────────────────────────────────────────────
case_files = sorted(os.listdir(JUDG_DIR))
results = []

print(f"\nProcessing {len(case_files)} cases...\n")
print(f"{'Case':<5} {'Case ID':<16} {'Doc Type':<24} {'Layman%':>9} {'BuddingLaw%':>12} {'Advocate%':>10}")
print("─" * 80)

DOC_ORDER = [
    ("Original Judgement",   JUDG_DIR,  0),
    ("Ground Truth Summary", GT_DIR,    1),
    ("LockNKey Summary",     LK_DIR,    2),
]

for i, fname in enumerate(case_files):
    case_num = i + 1
    case_id  = fname.replace(".txt", "")

    for doc_label, folder, offset in DOC_ORDER:
        text = open(os.path.join(folder, fname), encoding="utf-8", errors="ignore").read()
        lay, stu, pro = audience_scores(text)

        results.append({
            "case_num":             case_num,
            "case_id":              case_id,
            "document_type":        doc_label,
            "layman_score_pct":     lay,
            "budding_lawyer_pct":   stu,
            "advocate_pct":         pro,
            "excel_row":            sheet2_row(case_num, offset),
        })

        print(f"  {case_num:<4} {case_id:<16} {doc_label:<24} {lay:>9.4f} {stu:>12.4f} {pro:>10.4f}")

    print()

# ── WRITE INTO EXCEL SHEET 2 ───────────────────────────────────────
print(f"Writing into Excel: {XLSX_PATH}  →  Sheet: 'Audience-aware capability'")
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb["Audience-aware capability"]

for r in results:
    xrow = r["excel_row"]
    ws.cell(row=xrow, column=COL_LAYMAN).value   = r["layman_score_pct"]
    ws.cell(row=xrow, column=COL_STUDENT).value  = r["budding_lawyer_pct"]
    ws.cell(row=xrow, column=COL_ADVOCATE).value = r["advocate_pct"]

wb.save(XLSX_PATH)
print(f"  ✅ Sheet 2 filled — {len(results)} rows written ({len(case_files)} cases × 3 doc types)")

# ── SAVE CSV BACKUP ────────────────────────────────────────────────
with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=list(results[0].keys()))
    writer.writeheader()
    writer.writerows(results)

print(f"  ✅ CSV backup saved: {OUT_CSV}")

# ── SUMMARY ────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("AVERAGE SCORES ACROSS ALL 50 CASES")
print("=" * 60)
for doc_label, _, _ in DOC_ORDER:
    doc_rows = [r for r in results if r["document_type"] == doc_label]
    avg_lay  = round(sum(r["layman_score_pct"]   for r in doc_rows) / len(doc_rows), 4)
    avg_stu  = round(sum(r["budding_lawyer_pct"] for r in doc_rows) / len(doc_rows), 4)
    avg_pro  = round(sum(r["advocate_pct"]       for r in doc_rows) / len(doc_rows), 4)
    print(f"\n  {doc_label}:")
    print(f"    Avg Layman Score        : {avg_lay:.4f}%")
    print(f"    Avg Budding Lawyer Score: {avg_stu:.4f}%")
    print(f"    Avg Advocate Score      : {avg_pro:.4f}%")